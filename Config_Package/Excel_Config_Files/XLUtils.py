import time

import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def getColumnCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, colum_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=colum_no).value


def writeData(file, sheet_name, row_num, column_no, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # sheet.cell(row=row_num, column=column_no).value
    sheet.cell(row=row_num, column=column_no).value = str(data)
    workbook.save(file)


def result_pass(file, sheet_name, row_num, column_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row=row_num, column=column_no)
    cell.value = "PASS"
    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for pass
    workbook.save(file)


def result_fail(file, sheet_name, row_num, column_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row=row_num, column=column_no)
    cell.value = "FAIL"
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for fail
    workbook.save(file)


def totalTime(start_time, end_time):
    time_format = '%H:%M:%S'
    time1 = time.strptime(start_time, time_format)
    epoch_millis_start = time.mktime(time1)
    time2 = time.strptime(end_time, time_format)
    epoch_millis_end = time.mktime(time2)
    total = (epoch_millis_end - epoch_millis_start) / 1000
    return total


def getCurrentTime():
    now_time = time.strftime('%H:%M:%S')
    return now_time

